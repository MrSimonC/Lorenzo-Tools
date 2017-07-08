from custom_modules.file_to_db import FileToDB
from custom_modules.mssql import QueryDB
from custom_modules.xlsx import XlsxTools
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import RED, YELLOW
from openpyxl.formatting.rule import FormulaRule
import os

# Manually put the domain data into C:\Database files\Compare\input\Lorenzo Domain Data.xlsx

# Grab all dcs data
import lorenzo_sp_from_spreadsheets as lor_sp
folder = r'I:\Lorenzo Implementation\Documentation\System Configuration\Go-Live Build\Clinic Build'
results_file_output = r'C:\Database files\Compare\input\DCS Data.csv'
lor_sp.service_point_data_pull_all(folder, results_file_output)

# Process all files in input folder, import to database, run sql, output to excel
input_folder = r'C:\Database files\Compare\input'
pipes_folder = r'C:\Database files\Compare\pipes'
results_file = r'C:\Database files\Compare\results.xlsx'

files_to_process = [
    {'file_in_name': 'DCS Data.csv',
     'sheet_name': 'DCS Data',
     'table_name': 'dbo.DCS_Data'},
    # {'file_in_name': 'Lorenzo Domain Data.xlsx',
    #  'sheet_name': 'domain',
    #  'table_name': 'dbo.Lorenzo_Domain_Data'}
]

ftd = FileToDB('SERVER', 'DATABASE', 'USERNAME', 'PASSWORD')
for file in files_to_process:
    print("Loading: " + file['table_name'])
    ftd.import_file(os.path.join(input_folder, file['file_in_name']),
                    pipes_folder,
                    file['table_name'],
                    file['sheet_name'],
                    header_row_cell_value=file['header_row_cell_value'] if 'header_row_cell_value' in file else '')
print('Processing results')
db = QueryDB('SERVER', 'DATABASE', 'USERNAME', 'PASSWORD')
results = db.exec_sql('EXEC dbo.Lorenzo_DCS_to_Domain')
if results:
    XlsxTools().create_document(results, 'DCS vs Domain', results_file)
else:
    print('No results were returned')

print('Add conditional formatting')
wb = load_workbook(results_file)
ws = wb.active
ws.conditional_formatting.add('A:N', FormulaRule(formula=['$A1="Domain"'], font=Font(color=RED)))
ws.conditional_formatting.add('A:N', FormulaRule(formula=['$A1="DCS"'], font=Font(color='008000')))
pattern_yellow = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
ws.conditional_formatting.add('G:G', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1<>$G2)'], fill=pattern_yellow))
ws.conditional_formatting.add('H:H', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1=$G2,$H1<>$H2)'], fill=pattern_yellow))
ws.conditional_formatting.add('I:I', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1=$G2,$J1=$J2,$I1<>$I2)'], fill=pattern_yellow))
ws.conditional_formatting.add('J:J', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1=$G2,$K1=$K2,$J1<>$J2)'], fill=pattern_yellow))
ws.conditional_formatting.add('K:K', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1=$G2,$J1=$J2,$K1<>$K2)'], fill=pattern_yellow))
ws.conditional_formatting.add('L:L', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1=$G2,$J1=$J2,$L1<>$L2)'], fill=pattern_yellow))
ws.conditional_formatting.add('M:M', FormulaRule(formula=['AND($A1<>$A2,$E1=$E2,$G1=$G2,$J1=$J2,$M1<>$M2)'], fill=pattern_yellow))

wb.save(results_file)
